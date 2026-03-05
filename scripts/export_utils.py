#!/usr/bin/env python3
"""
导出工具 - 支持多种格式导出分镜脚本和角色设定
支持格式: JSON, CSV, Markdown, Excel, HTML

v2.0 安全改进:
- HTML 输出使用 html.escape 防止 XSS
- 文件名 slugify + 路径穿越校验
- CSV/Excel 公式注入防护（= + - @ 前缀净化）
"""

import json
import csv
import re
import html as html_module
from pathlib import Path
from typing import Dict, List, Optional, Any
from datetime import datetime


# ── 安全工具函数 ──────────────────────────────────────

def _sanitize_filename(name: str) -> str:
    """将标题转为安全文件名，防止路径穿越"""
    # 移除路径分隔符和危险字符
    safe = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', name)
    # 移除 .. 路径穿越
    safe = safe.replace('..', '_')
    # 限制长度
    return safe[:100] or 'untitled'


def _validate_output_path(file_path: Path, output_dir: Path):
    """校验输出路径必须在 output_dir 内"""
    resolved = file_path.resolve()
    dir_resolved = output_dir.resolve()
    if not str(resolved).startswith(str(dir_resolved)):
        raise ValueError(
            f"路径穿越检测: {file_path} 不在输出目录 {output_dir} 内"
        )


def _sanitize_cell(value: Any) -> Any:
    """CSV/Excel 公式注入防护：对以 = + - @ 开头的字符串加前缀"""
    if isinstance(value, str) and value and value[0] in ('=', '+', '-', '@'):
        return "'" + value
    return value


def _esc(text: Any) -> str:
    """HTML 转义"""
    return html_module.escape(str(text)) if text else ""


class ExportUtils:
    """导出工具类"""

    @staticmethod
    def export_to_json(data: Dict, output_path: str, indent: int = 2) -> str:
        """导出为 JSON 格式"""
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=indent)

        return str(path)

    @staticmethod
    def export_to_csv(shots: List[Dict], output_path: str) -> str:
        """导出分镜为 CSV 格式"""
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        if not shots:
            return str(path)

        # 定义字段顺序
        fieldnames = [
            'shot_id', 'scene_number', 'shot_number',
            'shot_size', 'camera_movement', 'subject',
            'action', 'dialogue', 'mood', 'duration',
            'transition', 'visual_prompt'
        ]

        with open(path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            # 公式注入防护
            for shot in shots:
                safe_row = {k: _sanitize_cell(v) for k, v in shot.items()}
                writer.writerow(safe_row)

        return str(path)

    @staticmethod
    def export_to_markdown(
        storyboard: Dict,
        characters: Dict[str, Dict],
        scenes: List[Dict],
        output_path: str
    ) -> str:
        """导出为 Markdown 格式的分镜脚本"""
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        lines = []

        # 标题
        title = storyboard.get('title', 'Untitled')
        lines.append(f"# {title} - 分镜脚本")
        lines.append("")
        lines.append(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("")

        # 元数据
        metadata = storyboard.get('metadata', {})
        lines.append("## 项目信息")
        lines.append("")
        lines.append(f"- 总场景数: {metadata.get('total_scenes', 0)}")
        lines.append(f"- 总镜头数: {metadata.get('total_shots', 0)}")
        lines.append(f"- 预估时长: {metadata.get('estimated_duration', 0):.1f} 秒")
        lines.append("")

        # 角色表
        lines.append("## 角色设定")
        lines.append("")
        for char_name, char_info in characters.items():
            lines.append(f"### {char_name}")
            lines.append("")
            lines.append(f"**视觉描述**: {char_info.get('prompt_description', 'N/A')}")
            lines.append("")
            keywords = char_info.get('visual_keywords', [])
            if keywords:
                lines.append(f"**关键词**: {', '.join(keywords)}")
                lines.append("")

        # 场景表
        lines.append("## 场景设定")
        lines.append("")
        for scene in scenes:
            scene_num = scene.get('scene_number', 0)
            env = scene.get('environment', {})
            lines.append(f"### 场景 {scene_num}: {env.get('location_type', 'Unknown')}")
            lines.append("")
            lines.append(f"- 类型: {env.get('int_ext', 'INT')}")
            lines.append(f"- 时间: {env.get('time_of_day', 'DAY')}")
            lines.append(f"- 视觉: {scene.get('visual_prompt', 'N/A')}")
            lines.append("")

        # 分镜表
        lines.append("## 分镜列表")
        lines.append("")

        # 按场景分组
        shots = storyboard.get('shots', [])
        current_scene = None

        for shot in shots:
            scene_num = shot.get('scene_number', 0)

            if scene_num != current_scene:
                current_scene = scene_num
                lines.append(f"### 场景 {scene_num}")
                lines.append("")
                lines.append("| 镜号 | 景别 | 运镜 | 主体 | 动作 | 时长 |")
                lines.append("|------|------|------|------|------|------|")

            lines.append(
                f"| {shot.get('shot_id', '')} "
                f"| {shot.get('shot_size', '')} "
                f"| {shot.get('camera_movement', '')} "
                f"| {shot.get('subject', '')} "
                f"| {shot.get('action', '')} "
                f"| {shot.get('duration', 0)}s |"
            )

        lines.append("")

        # 详细分镜
        lines.append("## 详细分镜提示词")
        lines.append("")

        for shot in shots:
            lines.append(f"### 镜头 {shot.get('shot_id', '')}")
            lines.append("")
            lines.append(f"**景别**: {shot.get('shot_size', '')}")
            lines.append(f"**运镜**: {shot.get('camera_movement', '')}")
            lines.append(f"**主体**: {shot.get('subject', '')}")
            lines.append(f"**动作**: {shot.get('action', '')}")
            if shot.get('dialogue'):
                lines.append(f"**对白**: {shot.get('dialogue', '')}")
            if shot.get('mood'):
                lines.append(f"**氛围**: {shot.get('mood', '')}")
            lines.append(f"**时长**: {shot.get('duration', 0)} 秒")
            lines.append(f"**转场**: {shot.get('transition', 'cut')}")
            lines.append("")
            lines.append("**AI生成提示词**:")
            lines.append(f"```")
            lines.append(shot.get('visual_prompt', ''))
            lines.append(f"```")
            lines.append("")
            lines.append("---")
            lines.append("")

        with open(path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))

        return str(path)

    @staticmethod
    def export_to_excel(
        storyboard: Dict,
        characters: Dict[str, Dict],
        scenes: List[Dict],
        output_path: str
    ) -> str:
        """导出为 Excel 格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise ImportError("请安装 openpyxl: pip install openpyxl")

        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()

        # 样式定义
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sheet 1: 分镜列表
        ws1 = wb.active
        ws1.title = "分镜列表"

        headers = ['镜号', '场景', '景别', '运镜', '主体', '动作', '对白', '氛围', '时长(秒)', '转场', '提示词']
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        shots = storyboard.get('shots', [])
        for row, shot in enumerate(shots, 2):
            data = [
                shot.get('shot_id', ''),
                shot.get('scene_number', ''),
                shot.get('shot_size', ''),
                shot.get('camera_movement', ''),
                shot.get('subject', ''),
                shot.get('action', ''),
                shot.get('dialogue', ''),
                shot.get('mood', ''),
                shot.get('duration', 0),
                shot.get('transition', ''),
                shot.get('visual_prompt', '')
            ]
            for col, value in enumerate(data, 1):
                cell = ws1.cell(row=row, column=col, value=_sanitize_cell(value))
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                cell.border = thin_border

        # 调整列宽
        ws1.column_dimensions['A'].width = 8
        ws1.column_dimensions['B'].width = 8
        ws1.column_dimensions['C'].width = 15
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 20
        ws1.column_dimensions['F'].width = 25
        ws1.column_dimensions['G'].width = 30
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 10
        ws1.column_dimensions['J'].width = 10
        ws1.column_dimensions['K'].width = 50

        # Sheet 2: 角色设定
        ws2 = wb.create_sheet("角色设定")
        char_headers = ['角色名', '视觉描述', '关键词', '出场场景']
        for col, header in enumerate(char_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for row, (name, info) in enumerate(characters.items(), 2):
            ws2.cell(row=row, column=1, value=name).border = thin_border
            ws2.cell(row=row, column=2, value=info.get('prompt_description', '')).border = thin_border
            ws2.cell(row=row, column=3, value=', '.join(info.get('visual_keywords', []))).border = thin_border
            ws2.cell(row=row, column=4, value=', '.join(map(str, info.get('scene_appearances', [])))).border = thin_border

        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 40
        ws2.column_dimensions['C'].width = 30
        ws2.column_dimensions['D'].width = 20

        # Sheet 3: 场景设定
        ws3 = wb.create_sheet("场景设定")
        scene_headers = ['场景号', '地点', '类型', '时间', '视觉描述', '氛围关键词']
        for col, header in enumerate(scene_headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for row, scene in enumerate(scenes, 2):
            env = scene.get('environment', {})
            ws3.cell(row=row, column=1, value=scene.get('scene_number', '')).border = thin_border
            ws3.cell(row=row, column=2, value=env.get('location_type', '')).border = thin_border
            ws3.cell(row=row, column=3, value=env.get('int_ext', '')).border = thin_border
            ws3.cell(row=row, column=4, value=env.get('time_of_day', '')).border = thin_border
            ws3.cell(row=row, column=5, value=scene.get('visual_prompt', '')).border = thin_border
            ws3.cell(row=row, column=6, value=', '.join(scene.get('mood_keywords', []))).border = thin_border

        ws3.column_dimensions['A'].width = 10
        ws3.column_dimensions['B'].width = 20
        ws3.column_dimensions['C'].width = 10
        ws3.column_dimensions['D'].width = 12
        ws3.column_dimensions['E'].width = 40
        ws3.column_dimensions['F'].width = 25

        wb.save(path)
        return str(path)

    @staticmethod
    def export_to_html(
        storyboard: Dict,
        characters: Dict[str, Dict],
        scenes: List[Dict],
        output_path: str
    ) -> str:
        """导出为可视化 HTML 格式"""
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        title = storyboard.get('title', 'Untitled')
        metadata = storyboard.get('metadata', {})
        shots = storyboard.get('shots', [])

        # 所有用户内容使用 _esc() 转义
        html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{_esc(title)} - 分镜脚本</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f5f5f5; color: #333; }}
        .container {{ max-width: 1200px; margin: 0 auto; padding: 20px; }}
        h1 {{ text-align: center; padding: 30px 0; color: #2c3e50; }}
        .meta {{ background: #fff; padding: 20px; border-radius: 8px; margin-bottom: 20px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        .meta-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; }}
        .meta-item {{ text-align: center; }}
        .meta-value {{ font-size: 24px; font-weight: bold; color: #3498db; }}
        .meta-label {{ color: #7f8c8d; font-size: 14px; }}
        .section {{ background: #fff; padding: 20px; border-radius: 8px; margin-bottom: 20px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        .section h2 {{ color: #2c3e50; margin-bottom: 15px; padding-bottom: 10px; border-bottom: 2px solid #3498db; }}
        .shot-grid {{ display: grid; gap: 15px; }}
        .shot-card {{ border: 1px solid #e0e0e0; border-radius: 8px; overflow: hidden; }}
        .shot-header {{ background: #3498db; color: #fff; padding: 10px 15px; display: flex; justify-content: space-between; align-items: center; }}
        .shot-id {{ font-weight: bold; }}
        .shot-duration {{ background: rgba(255,255,255,0.2); padding: 2px 8px; border-radius: 4px; font-size: 12px; }}
        .shot-body {{ padding: 15px; }}
        .shot-info {{ display: grid; grid-template-columns: repeat(2, 1fr); gap: 10px; margin-bottom: 10px; }}
        .shot-info-item {{ }}
        .shot-info-label {{ font-size: 12px; color: #7f8c8d; }}
        .shot-info-value {{ font-weight: 500; }}
        .shot-prompt {{ background: #f8f9fa; padding: 10px; border-radius: 4px; font-family: monospace; font-size: 12px; white-space: pre-wrap; word-break: break-all; }}
        .character-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 15px; }}
        .character-card {{ border: 1px solid #e0e0e0; border-radius: 8px; padding: 15px; }}
        .character-name {{ font-size: 18px; font-weight: bold; color: #2c3e50; margin-bottom: 10px; }}
        .character-desc {{ color: #555; margin-bottom: 10px; }}
        .keywords {{ display: flex; flex-wrap: wrap; gap: 5px; }}
        .keyword {{ background: #e8f4f8; color: #3498db; padding: 2px 8px; border-radius: 4px; font-size: 12px; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{_esc(title)}</h1>

        <div class="meta">
            <div class="meta-grid">
                <div class="meta-item">
                    <div class="meta-value">{metadata.get('total_scenes', 0)}</div>
                    <div class="meta-label">场景数</div>
                </div>
                <div class="meta-item">
                    <div class="meta-value">{metadata.get('total_shots', 0)}</div>
                    <div class="meta-label">镜头数</div>
                </div>
                <div class="meta-item">
                    <div class="meta-value">{metadata.get('estimated_duration', 0):.0f}s</div>
                    <div class="meta-label">预估时长</div>
                </div>
                <div class="meta-item">
                    <div class="meta-value">{len(characters)}</div>
                    <div class="meta-label">角色数</div>
                </div>
            </div>
        </div>

        <div class="section">
            <h2>👤 角色设定</h2>
            <div class="character-grid">
"""

        for name, info in characters.items():
            keywords_html = ''.join(f'<span class="keyword">{_esc(kw)}</span>' for kw in info.get('visual_keywords', [])[:5])
            html += f"""
                <div class="character-card">
                    <div class="character-name">{_esc(name)}</div>
                    <div class="character-desc">{_esc(info.get('prompt_description', 'N/A'))}</div>
                    <div class="keywords">{keywords_html}</div>
                </div>
"""

        html += """
            </div>
        </div>

        <div class="section">
            <h2>🎬 分镜列表</h2>
            <div class="shot-grid">
"""

        for shot in shots:
            html += f"""
                <div class="shot-card">
                    <div class="shot-header">
                        <span class="shot-id">镜头 {_esc(shot.get('shot_id', ''))}</span>
                        <span class="shot-duration">{_esc(shot.get('duration', 0))}s</span>
                    </div>
                    <div class="shot-body">
                        <div class="shot-info">
                            <div class="shot-info-item">
                                <div class="shot-info-label">景别</div>
                                <div class="shot-info-value">{_esc(shot.get('shot_size', ''))}</div>
                            </div>
                            <div class="shot-info-item">
                                <div class="shot-info-label">运镜</div>
                                <div class="shot-info-value">{_esc(shot.get('camera_movement', ''))}</div>
                            </div>
                            <div class="shot-info-item">
                                <div class="shot-info-label">主体</div>
                                <div class="shot-info-value">{_esc(shot.get('subject', ''))}</div>
                            </div>
                            <div class="shot-info-item">
                                <div class="shot-info-label">动作</div>
                                <div class="shot-info-value">{_esc(shot.get('action', ''))}</div>
                            </div>
                        </div>
                        <div class="shot-prompt">{_esc(shot.get('visual_prompt', ''))}</div>
                    </div>
                </div>
"""

        html += """
            </div>
        </div>
    </div>
</body>
</html>
"""

        with open(path, 'w', encoding='utf-8') as f:
            f.write(html)

        return str(path)


def export_all(
    storyboard: Dict,
    characters: Dict[str, Dict],
    scenes: List[Dict],
    output_dir: str,
    formats: List[str] = None
) -> Dict[str, str]:
    """
    批量导出所有格式

    Args:
        storyboard: 分镜数据
        characters: 角色数据
        scenes: 场景数据
        output_dir: 输出目录
        formats: 要导出的格式列表,默认全部

    Returns:
        格式 -> 文件路径 的映射
    """
    if formats is None:
        formats = ['json', 'csv', 'markdown', 'excel', 'html']

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    title = _sanitize_filename(storyboard.get('title', 'storyboard'))
    results = {}

    if 'json' in formats:
        json_path = output_dir / f"{title}.json"
        _validate_output_path(json_path, output_dir)
        results['json'] = ExportUtils.export_to_json(
            {'storyboard': storyboard, 'characters': characters, 'scenes': scenes},
            str(json_path)
        )

    if 'csv' in formats:
        csv_path = output_dir / f"{title}.csv"
        _validate_output_path(csv_path, output_dir)
        results['csv'] = ExportUtils.export_to_csv(
            storyboard.get('shots', []),
            str(csv_path)
        )

    if 'markdown' in formats:
        md_path = output_dir / f"{title}.md"
        _validate_output_path(md_path, output_dir)
        results['markdown'] = ExportUtils.export_to_markdown(
            storyboard, characters, scenes,
            str(md_path)
        )

    if 'excel' in formats:
        try:
            xlsx_path = output_dir / f"{title}.xlsx"
            _validate_output_path(xlsx_path, output_dir)
            results['excel'] = ExportUtils.export_to_excel(
                storyboard, characters, scenes,
                str(xlsx_path)
            )
        except ImportError as e:
            print(f"Excel导出跳过: {e}")

    if 'html' in formats:
        html_path = output_dir / f"{title}.html"
        _validate_output_path(html_path, output_dir)
        results['html'] = ExportUtils.export_to_html(
            storyboard, characters, scenes,
            str(html_path)
        )

    return results


if __name__ == "__main__":
    print("Export utilities loaded. Use export_all() for batch export.")
